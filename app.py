import os
import sys
import time
from threading import Timer
import webbrowser
import calendar 

# --- MARCADOR VISUAL ---
print("\n" + "="*60)
print(">>> VERSÃO 5.1: CORREÇÃO DO ERRO JSON NO EDITAR <<<")
print("="*60 + "\n")

# --- TESTE DE BIBLIOTECAS ---
try:
    import flask
    import pandas as pd
    import openpyxl
    import sqlite3
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    print(">>> BIBLIOTECAS OK.")
except ImportError as e:
    print(f"\n>>> ERRO: Falta {e.name}. Rode: pip install {e.name} pillow\n")
    sys.exit(1)

from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash
from io import BytesIO
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'transer_segredo_total'
DB_NAME = 'transer.db'

# Configurações
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'temp_uploads')
if not os.path.exists(UPLOAD_FOLDER): os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def init_db():
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS clients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                external_id TEXT,
                document TEXT,
                address TEXT,
                city TEXT,
                email TEXT,
                contract_num TEXT,
                contract_val REAL,
                contract_limit REAL,
                extra_val REAL,
                periodicity TEXT,
                created_at TEXT,
                type TEXT DEFAULT 'standard',
                parent_id INTEGER,
                price_kg REAL
            )
        ''')
        cursor.execute('''CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE, password TEXT, role TEXT, name TEXT)''')
        
        cursor.execute("SELECT * FROM users WHERE username = 'admin'")
        if not cursor.fetchone():
            cursor.execute("INSERT INTO users (username, password, role, name) VALUES (?, ?, ?, ?)", ('admin', 'admambiental', 'admin', 'Admin Padrão'))
        conn.commit()

# --- ROTAS BÁSICAS ---
@app.route('/')
def index():
    if 'user_id' not in session: return redirect(url_for('login'))
    return render_template('index.html', view='dashboard', user=session.get('user_name'), role=session.get('role'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u, p = request.form['username'], request.form['password']
        with sqlite3.connect(DB_NAME) as conn:
            user = conn.execute("SELECT * FROM users WHERE username = ? AND password = ?", (u, p)).fetchone()
            if user:
                session['user_id'], session['role'], session['user_name'] = user[0], user[3], user[4]
                return redirect(url_for('index'))
            else: flash('Login inválido', 'error')
    return render_template('index.html', view='login')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# --- GESTÃO DE CLIENTES ---
@app.route('/clients', methods=['GET', 'POST'])
def clients():
    if 'user_id' not in session: return redirect(url_for('login'))
    
    if request.method == 'POST':
        d = request.form
        try:
            with sqlite3.connect(DB_NAME) as conn:
                now = datetime.now().strftime("%Y-%m-%d")
                pid = d.get('parent_id') if d.get('parent_id') else None
                pkg = d.get('price_kg') if d.get('price_kg') else 0
                cval = d.get('contract_val') if d.get('contract_val') else 0
                clim = d.get('contract_limit') if d.get('contract_limit') else 0
                cxtra = d.get('extra_val') if d.get('extra_val') else 0
                ext_id = d.get('external_id') if d.get('external_id') else None 

                if d.get('client_id'):
                    conn.execute('''
                        UPDATE clients SET name=?, external_id=?, document=?, address=?, city=?, email=?, 
                        contract_num=?, contract_val=?, contract_limit=?, extra_val=?, periodicity=?, 
                        type=?, parent_id=?, price_kg=? WHERE id=?
                    ''', (d['name'], ext_id, d['document'], d['address'], d['city'], d['email'], 
                          d['contract_num'], cval, clim, cxtra, d['periodicity'], 
                          d['type'], pid, pkg, d['client_id']))
                    flash('Cliente atualizado com sucesso!', 'success')
                else:
                    conn.execute('''
                        INSERT INTO clients (name, external_id, document, address, city, email, 
                        contract_num, contract_val, contract_limit, extra_val, periodicity, created_at,
                        type, parent_id, price_kg) 
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    ''', (d['name'], ext_id, d['document'], d['address'], d['city'], d['email'], 
                          d['contract_num'], cval, clim, cxtra, d['periodicity'], now,
                          d['type'], pid, pkg))
                    flash('Cliente cadastrado com sucesso!', 'success')
        except Exception as e: 
            flash(f'Erro: {e}', 'error')
            print(e)

    with sqlite3.connect(DB_NAME) as conn:
        conn.row_factory = sqlite3.Row
        # AQUI ESTAVA O PROBLEMA: Converter Row para Dict para o JSON funcionar
        clients_rows = conn.execute("SELECT * FROM clients ORDER BY name").fetchall()
        clients = [dict(row) for row in clients_rows]
        
        parents_rows = conn.execute("SELECT * FROM clients WHERE type = 'master' ORDER BY name").fetchall()
        parents = [dict(row) for row in parents_rows]
        
    return render_template('index.html', view='clients', clients=clients, parents=parents, user=session.get('user_name'), role=session.get('role'))

@app.route('/delete_client/<int:id>')
def delete_client(id):
    if 'user_id' not in session: return redirect(url_for('login'))
    with sqlite3.connect(DB_NAME) as conn: conn.execute("DELETE FROM clients WHERE id=?", (id,))
    return redirect(url_for('clients'))

# --- GESTÃO DE USUÁRIOS ---
@app.route('/users', methods=['GET', 'POST'])
def users():
    if 'user_id' not in session or session.get('role') != 'admin': return redirect(url_for('index'))
    if request.method == 'POST':
        d = request.form
        try:
            with sqlite3.connect(DB_NAME) as conn: conn.execute("INSERT INTO users (name, username, password, role) VALUES (?,?,?,?)", (d['name'], d['username'], d['password'], d['role']))
            flash('Usuário criado!', 'success')
        except: flash('Login já existe', 'error')
    with sqlite3.connect(DB_NAME) as conn:
        conn.row_factory = sqlite3.Row
        return render_template('index.html', view='users', users=conn.execute("SELECT * FROM users").fetchall(), user=session.get('user_name'), role=session.get('role'))

@app.route('/delete_user/<int:id>')
def delete_user(id):
    if 'user_id' not in session or session.get('role') != 'admin': return redirect(url_for('index'))
    if id != session.get('user_id'): 
        with sqlite3.connect(DB_NAME) as conn: conn.execute("DELETE FROM users WHERE id=?", (id,))
    return redirect(url_for('users'))

@app.route('/reports')
def reports():
    if 'user_id' not in session: return redirect(url_for('login'))
    with sqlite3.connect(DB_NAME) as conn:
        conn.row_factory = sqlite3.Row
        clients = conn.execute("SELECT * FROM clients ORDER BY id DESC").fetchall()
        stats = {'total': len(clients), 'val': sum(c['contract_val'] or 0 for c in clients)}
    return render_template('index.html', view='reports', clients=clients, stats=stats, user=session.get('user_name'), role=session.get('role'))

# --- FECHAMENTO INTELIGENTE ---
@app.route('/closing', methods=['GET', 'POST'])
def closing():
    if 'user_id' not in session: return redirect(url_for('login'))
    
    billable_items = {} 
    
    if request.method == 'POST' and 'file' in request.files:
        file = request.files['file']
        if file.filename:
            ext = ".xlsx" if file.filename.lower().endswith(('.xlsx', '.xls')) else ".csv"
            temp_filename = f"temp_import_{int(datetime.now().timestamp())}{ext}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)
            try:
                file.save(filepath)
                df = None
                
                if ext == ".xlsx":
                    df = pd.read_excel(filepath, engine='openpyxl')
                    cols = [str(c).strip() for c in df.columns]
                    if 'ID Cliente' not in cols: df = pd.read_excel(filepath, engine='openpyxl', header=1)
                else:
                    try: df = pd.read_csv(filepath, encoding='latin1', sep=';', on_bad_lines='skip')
                    except: df = pd.read_csv(filepath, encoding='latin1', sep=',', on_bad_lines='skip')
                    cols = [str(c).strip() for c in df.columns]
                    if 'ID Cliente' not in cols:
                         try: df = pd.read_csv(filepath, encoding='latin1', sep=';', on_bad_lines='skip', header=1)
                         except: df = pd.read_csv(filepath, encoding='latin1', sep=',', on_bad_lines='skip', header=1)

                if df is not None:
                    df.columns = [str(c).strip() for c in df.columns]
                    if 'ID Cliente' not in df.columns:
                        flash('Erro: Coluna "ID Cliente" não encontrada.', 'error')
                    else:
                        with sqlite3.connect(DB_NAME) as conn:
                            conn.row_factory = sqlite3.Row
                            db_clients = conn.execute("SELECT * FROM clients").fetchall()
                        
                        map_by_ext_id = {str(c['external_id']): c for c in db_clients if c['external_id']}
                        map_by_db_id = {c['id']: c for c in db_clients}

                        for cid_raw, group in df.groupby('ID Cliente'):
                            cid = str(cid_raw).replace('.0', '')
                            client_info = map_by_ext_id.get(cid)
                            
                            if not client_info:
                                master_id = f"UNK_{cid}"
                                master_data = {'name': f"Desconhecido (ID: {cid})", 'id': cid, 'type': 'unknown'}
                                display_name = f"Desconhecido (ID: {cid})"
                            else:
                                if client_info['type'] == 'child' and client_info['parent_id']:
                                    # Filial -> Joga na Matriz
                                    parent = map_by_db_id.get(client_info['parent_id'])
                                    if parent:
                                        master_id = str(parent['id'])
                                        master_data = dict(parent)
                                        display_name = client_info['name']
                                    else:
                                        master_id = str(client_info['id'])
                                        master_data = dict(client_info)
                                        display_name = client_info['name']
                                else:
                                    master_id = str(client_info['id'])
                                    master_data = dict(client_info)
                                    display_name = client_info['name']

                            if master_id not in billable_items:
                                billable_items[master_id] = {
                                    'client_db': master_data,
                                    'items': [],
                                    'total_kg': 0,
                                    'status': 'OK' if client_info else 'Off'
                                }

                            for idx, row in group.iterrows():
                                qty = str(row.get('Quantidade', 0)).replace(',', '.')
                                try: qty = float(qty)
                                except: qty = 0
                                
                                billable_items[master_id]['items'].append({
                                    'display_name': display_name,
                                    'date': row.get('Data'),
                                    'plate': row.get('Placa do Veículo', '-'),
                                    'waste_class': row.get('Classe do Resíduo'),
                                    'qty': qty
                                })
                                billable_items[master_id]['total_kg'] += qty

                else: flash("Arquivo ilegível.", 'error')
            except Exception as e:
                print(f"ERRO: {e}"); flash(f'Erro técnico: {e}', 'error')
            finally:
                if os.path.exists(filepath):
                    try: os.remove(filepath)
                    except: pass
    
    # Cálculos Finais com Lógica de Período (Trimestral/Semestral)
    final_list = []
    multipliers = {'Mensal': 1, 'Trimestral': 3, 'Semestral': 6, 'Anual': 12, 'Ocasional': 1}

    for key, data in billable_items.items():
        c = data['client_db']
        total_kg = data['total_kg']
        client_type = c.get('type')
        periodicity = c.get('periodicity', 'Mensal')
        
        # Pega o multiplicador (Padrão é 1)
        factor = multipliers.get(periodicity, 1)
        # Se for Ocasional ou não padrão, assume 1 para evitar sustos, a menos que especificado
        if periodicity == 'Ocasional': factor = 1 

        if client_type == 'master' or client_type == 'kg_private':
            # PREFEITURA OU PARTICULAR POR KG
            price = c.get('price_kg', 0)
            total_val = total_kg * price
            excess = 0 
        else: 
            # PADRÃO: Pacote
            # Aqui aplicamos a multiplicação pelo período
            limit_base = c.get('contract_limit', 0)
            pkg_val_base = c.get('contract_val', 0)
            
            limit = limit_base * factor
            pkg_val = pkg_val_base * factor
            
            extra_val = c.get('extra_val', 0)
            excess = max(0, total_kg - limit)
            total_val = pkg_val + (excess * extra_val)
            
        data['total_val'] = total_val
        data['excess'] = excess
        final_list.append(data)

    return render_template('index.html', view='closing', processed=final_list, user=session.get('user_name'), role=session.get('role'))

# --- GERAÇÃO DO EXCEL ---
@app.route('/generate_excel', methods=['POST'])
def generate_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.drawing.image import Image
    except ImportError: return "Erro: openpyxl não instalado.", 500

    req_data = request.json
    c = req_data['client']
    items = req_data['coletas']
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    
    # Estilos
    font_bold = Font(name='Arial', size=10, bold=True)
    font_reg = Font(name='Arial', size=10)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fmt_money = '#,##0.00'
    fmt_dec = '0.00'
    
    # Logo
    logo_path = os.path.join(BASE_DIR, 'logo.png')
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            max_w, max_h = 250, 100
            ratio = min(max_w/img.width, max_h/img.height, 1)
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
            ws.add_image(img, 'A1')
        except: pass

    # Cálculos de Data (Min/Max do arquivo)
    dates = []
    for x in items:
        try: dates.append(datetime.strptime(str(x.get('date')).split()[0], '%Y-%m-%d'))
        except: pass
    
    if dates:
        min_date = min(dates).strftime('%d/%m/%Y')
        max_date = max(dates).strftime('%d/%m/%Y')
        periodo_real = f"{min_date} A {max_date}"
    else:
        # Fallback se não achar datas
        now = datetime.now()
        last_day = calendar.monthrange(now.year, now.month)[1]
        periodo_real = f"01/{now.month:02d}/{now.year} A {last_day}/{now.month:02d}/{now.year}"

    meses = {1: 'JANEIRO', 2: 'FEVEREIRO', 3: 'MARÇO', 4: 'ABRIL', 5: 'MAIO', 6: 'JUNHO', 7: 'JULHO', 8: 'AGOSTO', 9: 'SETEMBRO', 10: 'OUTUBRO', 11: 'NOVEMBRO', 12: 'DEZEMBRO'}
    now = datetime.now()
    mes_ano = f"{meses[now.month]} DE {now.year}"
    
    row_num = 7
    ws.cell(row=row_num, column=1, value=f"FECHAMENTO {mes_ano}").font = Font(name='Arial', size=12, bold=True)
    row_num += 1
    ws.cell(row=row_num, column=1, value="TRANSER AMBIENTAL LTDA").font = font_bold
    row_num += 1
    
    fields = [("CLIENTE:", c['name']), ("E-MAIL:", c['email']), ("ENDEREÇO:", c['address']), ("CIDADE:", c['city']), ("CPF/CNPJ:", c['document']), ("Período:", periodo_real), ("CONTRATO Nº:", c['contract_num'])]
    for k, v in fields:
        ws.cell(row=row_num, column=1, value=k).font = font_bold
        ws.cell(row=row_num, column=2, value=v).font = font_reg
        row_num += 1
    row_num += 2

    client_type = c.get('type')

    if client_type == 'master': 
        # >>> LAYOUT 1: PREFEITURA (TABELA COMPLEXA) <<<
        headers = ["DATA", "CLIENTE", "PLACA", "GRUPO", "PESO (Kg)", "VALOR DO Kg", "VALOR"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=i, value=h)
            cell.font = font_bold; cell.border = border_thin; cell.alignment = Alignment(horizontal='center')
        row_num += 1
        
        price_kg = c.get('price_kg', 0)
        total_kg = 0
        total_val = 0
        
        for item in items:
            dt = item.get('date')
            try: dt = datetime.strptime(str(dt).split()[0], '%Y-%m-%d').strftime('%d/%m/%Y')
            except: pass
            
            qty = float(item.get('qty', 0))
            val = qty * price_kg
            total_kg += qty
            total_val += val
            
            vals = [dt, item.get('display_name'), item.get('plate'), item.get('waste_class'), qty, price_kg, val]
            for i, v in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=i, value=v)
                cell.font = font_reg; cell.border = border_thin
                cell.alignment = Alignment(horizontal='center') if i != 2 else Alignment(horizontal='left')
                if i == 5: cell.number_format = fmt_dec 
                if i >= 6: cell.number_format = fmt_money 
            row_num += 1
            
        row_num += 1
        ws.cell(row=row_num, column=4, value="TOTAIS").font = font_bold
        ws.cell(row=row_num, column=5, value=total_kg).font = font_bold; ws.cell(row=row_num, column=5).border = border_thin; ws.cell(row=row_num, column=5).number_format = fmt_dec
        ws.cell(row=row_num, column=7, value=total_val).font = font_bold; ws.cell(row=row_num, column=7).border = border_thin; ws.cell(row=row_num, column=7).number_format = fmt_money
        
        ws.column_dimensions['B'].width = 40

    else:
        # >>> LAYOUT 2: PADRÃO ou PARTICULAR KG <<<
        
        headers = ["Data da Coleta", "Resíduo", "Qtd. KG", "Normal/Extra", "OBS."]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=i, value=h)
            cell.font = font_bold; cell.border = border_thin; cell.alignment = Alignment(horizontal='center')
        row_num += 1
        
        total_kg = 0
        for item in items:
            qty = float(item.get('qty', 0))
            total_kg += qty
            dt = item.get('date')
            try: dt = datetime.strptime(str(dt).split()[0], '%Y-%m-%d').strftime('%d/%m/%Y')
            except: pass
            
            ws.cell(row=row_num, column=1, value=dt).border = border_thin; ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=2, value=item.get('waste_class')).border = border_thin
            ws.cell(row=row_num, column=3, value=qty).border = border_thin; ws.cell(row=row_num, column=3).number_format = fmt_dec; ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=4, value="Normal").border = border_thin; ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=5, value="").border = border_thin
            row_num += 1
            
        ws.cell(row=row_num, column=1, value="Total").font = font_bold
        ws.cell(row=row_num, column=3, value=total_kg).font = font_bold; ws.cell(row=row_num, column=3).number_format = fmt_dec
        row_num += 2
        
        fin_h = [(1,"Descrição"), (3,"Quant."), (4,"Valor Unitário"), (5,"Valor Total")]
        for col, val in fin_h:
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = font_bold; cell.border = border_thin; cell.alignment = Alignment(horizontal='center')
        row_num += 1
        
        if client_type == 'kg_private':
            # FINANCEIRO: SIMPLES
            price_kg = c.get('price_kg', 0)
            total_final = total_kg * price_kg
            
            ws.cell(row=row_num, column=1, value="Resíduos Coletados").border = border_thin
            ws.cell(row=row_num, column=3, value=total_kg).border = border_thin; ws.cell(row=row_num, column=3).number_format = fmt_dec
            ws.cell(row=row_num, column=4, value=price_kg).border = border_thin; ws.cell(row=row_num, column=4).number_format = fmt_money
            ws.cell(row=row_num, column=5, value=total_final).border = border_thin; ws.cell(row=row_num, column=5).number_format = fmt_money
            row_num += 1
            
        else:
            # FINANCEIRO: PADRÃO (Com Multiplicador de Período)
            multipliers = {'Mensal': 1, 'Trimestral': 3, 'Semestral': 6, 'Anual': 12, 'Ocasional': 1}
            periodicity = c.get('periodicity', 'Mensal')
            factor = multipliers.get(periodicity, 1)
            if periodicity == 'Ocasional': factor = 1

            limit_base = c.get('contract_limit', 0)
            pkg_val_base = c.get('contract_val', 0)
            
            limit = limit_base * factor
            pkg_val = pkg_val_base * factor
            
            extra_price = c.get('extra_val', 0)
            excess = max(0, total_kg - limit)
            extra_cost = excess * extra_price
            total_final = pkg_val + extra_cost
            
            # Linha Pacote
            desc_pacote = "Pacote contratado"
            if factor > 1: desc_pacote += f" ({periodicity} - {factor}x)"
            
            ws.cell(row=row_num, column=1, value=desc_pacote).border = border_thin
            ws.cell(row=row_num, column=3, value=1).border = border_thin
            ws.cell(row=row_num, column=4, value=pkg_val).border = border_thin; ws.cell(row=row_num, column=4).number_format = fmt_money
            ws.cell(row=row_num, column=5, value=pkg_val).border = border_thin; ws.cell(row=row_num, column=5).number_format = fmt_money
            row_num += 1
            
            # Linha Excedente
            ws.cell(row=row_num, column=1, value=f"Excedente ({limit}Kg)").border = border_thin
            ws.cell(row=row_num, column=3, value=excess).border = border_thin; ws.cell(row=row_num, column=3).number_format = fmt_dec
            ws.cell(row=row_num, column=4, value=extra_price).border = border_thin; ws.cell(row=row_num, column=4).number_format = fmt_money
            ws.cell(row=row_num, column=5, value=extra_cost).border = border_thin; ws.cell(row=row_num, column=5).number_format = fmt_money
            row_num += 1
            
        ws.cell(row=row_num, column=1, value="Valor Total").font = font_bold; ws.cell(row=row_num, column=1).border = border_thin
        c_fin = ws.cell(row=row_num, column=5, value=total_final)
        c_fin.font = font_bold; c_fin.border = border_thin; c_fin.number_format = fmt_money
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

    out = BytesIO()
    wb.save(out); out.seek(0)
    return send_file(out, download_name=f"Fechamento_{c['name']}.xlsx", as_attachment=True)

if __name__ == '__main__':
    init_db()
    Timer(1, lambda: webbrowser.open_new('http://127.0.0.1:5000/')).start()
    app.run(debug=True, use_reloader=False)